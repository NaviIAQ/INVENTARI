from flask import Flask, render_template, request, jsonify, send_file
import json, os, copy
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl import load_workbook

app = Flask(__name__)

EXCEL_FILE = "Formulario_Inventario_Transfusiones.xlsx"

# Cargar productos del Excel al arrancar
def cargar_productos():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    productos = {}
    nomes_anotar = False  # Flag para sección "NOMÉS ANOTAR EXISTÈNCIES"
    for row in ws.iter_rows(values_only=True):
        codi = row[0]
        nom  = row[1]
        # Detectar sección NOMÉS ANOTAR
        if codi and 'ANOTAR' in str(codi).upper():
            nomes_anotar = True
            continue
        if codi is None or nom is None:
            continue
        # Saltar cabeceras y filas de sección
        codi_str = str(codi).strip()
        if codi_str in ('CODI', 'HCCG/HAGR/HBH01', 'HCFA/HAFA/HBH05',
                        'HCFA/HASE/HBH05', 'DATA:'):
            continue
        try:
            float(codi_str.replace('.0',''))
        except:
            continue
        codi_clean = codi_str.replace('.0', '') if codi_str.endswith('.0') else codi_str
        productos[codi_clean] = {
            "codi": codi_clean,
            "nom": str(nom).strip(),
            "unitat": str(row[2]).strip() if row[2] else "",
            "estoc_ideal": str(row[3]).strip() if row[3] else "",
            "nomes_anotar": nomes_anotar,
            "comanda": None,
            "hora": None,
            "treballador": None,
        }
    return productos

# Stock en memoria (se resetea al reiniciar)
stock = {}

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/qr")
def qr():
    import qrcode
    url = os.environ.get('RENDER_EXTERNAL_URL', request.host_url).rstrip('/')
    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

@app.route("/api/producte/<codi>")
def get_producte(codi):
    codi = codi.replace('.0','')
    productes = cargar_productos()
    p = productes.get(codi)
    if p:
        # Añadir si ya tiene comanda registrada
        if codi in stock:
            p["comanda"] = stock[codi]["comanda"]
            p["hora"] = stock[codi]["hora"]
        return jsonify({"trobat": True, **p})
    return jsonify({"trobat": False, "codi": codi})

@app.route("/api/guardar", methods=["POST"])
def guardar():
    data = request.json
    codi = str(data["codi"]).replace('.0','')
    stock[codi] = {
        "codi": codi,
        "nom": data["nom"],
        "unitat": data.get("unitat",""),
        "estoc_ideal": data.get("estoc_ideal",""),
        "comanda": data["comanda"],
        "treballador": data.get("treballador","Anònim"),
        "hora": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    return jsonify({"ok": True})

@app.route("/api/stock")
def get_stock():
    return jsonify(list(stock.values()))

@app.route("/api/resetear", methods=["POST"])
def resetear():
    stock.clear()
    return jsonify({"ok": True})

@app.route("/api/excel")
def descargar_excel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Fecha en celda DATA
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "DATA:":
                # Poner fecha en la celda de al lado
                ws.cell(row=cell.row, column=cell.column + 1,
                        value=datetime.now().strftime("%d/%m/%Y"))
                break

    # Rellenar solo COMANDA (columna E = 5)
    for row in ws.iter_rows():
        codi_cell = row[0]  # Columna A
        if codi_cell.value is None:
            continue
        codi_str = str(codi_cell.value).replace('.0','').strip()
        if codi_str in stock:
            entrada = stock[codi_str]
            ws.cell(row=codi_cell.row, column=5, value=entrada["comanda"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Inventari_Transfusions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

if __name__ == "__main__":
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
    except:
        ip = "127.0.0.1"
    print(f"\n{'='*45}")
    print(f"  🏥  INVENTARI BST — TRANSFUSIONS")
    print(f"{'='*45}")
    print(f"  App:   http://{ip}:8080/")
    print(f"  Local: http://127.0.0.1:8080/")
    print(f"{'='*45}\n")
    app.run(host="0.0.0.0", port=8080, debug=True)
