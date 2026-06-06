from flask import Flask, render_template, request, jsonify, send_file
import json, os
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook

app = Flask(__name__)

EXCEL_FILES = {
    "transfusions": "Formulario_Inventario_Transfusiones.xlsx",
    "immuno": "Formulario_Inventario_Immuno.xlsx",
}

def cargar_productos(servei):
    excel_file = EXCEL_FILES.get(servei)
    if not excel_file or not os.path.exists(excel_file):
        return {}
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active
    productos = {}
    nomes_anotar = False
    for row in ws.iter_rows(values_only=True):
        codi = row[0]
        nom  = row[1]
        if codi and 'ANOTAR' in str(codi).upper():
            nomes_anotar = True
            continue
        if codi is None or nom is None:
            continue
        codi_str = str(codi).strip()
        if codi_str in ('CODI', 'HCCG/HAGR/HBH01', 'HCFA/HAFA/HBH05',
                        'HCFA/HASE/HBH05', 'DATA:', 'IMMUNO', 'TRANSFUSIONS'):
            continue
        try:
            float(codi_str.replace('.0',''))
        except:
            continue
        codi_clean = codi_str.replace('.0', '') if codi_str.endswith('.0') else codi_str
        productos[codi_clean] = {
            "codi": codi_clean,
            "nom": str(nom).strip(),
            "unitat": str(row[2]).strip() if row[2] else "",
            "estoc_ideal": str(row[3]).strip() if row[3] else "",
            "nomes_anotar": nomes_anotar,
            "comanda": None,
            "hora": None,
            "treballador": None,
        }
    return productos

# Stock en memoria per servei
stock = {"transfusions": {}, "immuno": {}}

@app.route("/")
def index():
    return render_template("selector.html")

@app.route("/inventari/<servei>")
def inventari(servei):
    if servei not in EXCEL_FILES:
        return "Servei no trobat", 404
    return render_template("index.html", servei=servei)

@app.route("/qr")
def qr():
    import qrcode
    url = os.environ.get('RENDER_EXTERNAL_URL', request.host_url).rstrip('/')
    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

@app.route("/api/<servei>/producte/<codi>")
def get_producte(servei, codi):
    codi = codi.replace('.0','')
    productes = cargar_productos(servei)
    p = productes.get(codi)
    if p:
        if codi in stock.get(servei, {}):
            p["comanda"] = stock[servei][codi]["comanda"]
            p["hora"] = stock[servei][codi]["hora"]
        return jsonify({"trobat": True, **p})
    return jsonify({"trobat": False, "codi": codi})

@app.route("/api/<servei>/guardar", methods=["POST"])
def guardar(servei):
    data = request.json
    codi = str(data["codi"]).replace('.0','')
    if servei not in stock:
        stock[servei] = {}
    stock[servei][codi] = {
        "codi": codi,
        "nom": data["nom"],
        "unitat": data.get("unitat",""),
        "estoc_ideal": data.get("estoc_ideal",""),
        "comanda": data["comanda"],
        "treballador": data.get("treballador","Anònim"),
        "hora": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    return jsonify({"ok": True})

@app.route("/api/<servei>/total_productes")
def total_productes(servei):
    productes = cargar_productos(servei)
    return jsonify({"total": len(productes)})

@app.route("/api/<servei>/tots_productes")
def tots_productes(servei):
    productes = cargar_productos(servei)
    return jsonify(list(productes.values()))

@app.route("/api/<servei>/stock")
def get_stock(servei):
    return jsonify(list(stock.get(servei, {}).values()))

@app.route("/api/<servei>/resetear", methods=["POST"])
def resetear(servei):
    stock[servei] = {}
    return jsonify({"ok": True})

@app.route("/api/<servei>/excel")
def descargar_excel(servei):
    from openpyxl.utils import get_column_letter
    excel_file = EXCEL_FILES.get(servei)
    if not excel_file:
        return "Servei no trobat", 404
    wb = load_workbook(excel_file)
    ws = wb.active

    # Format DIN A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = '1:3'  # Repetir capçaleres a cada pàgina

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "DATA:":
                ws.cell(row=cell.row, column=cell.column + 1,
                        value=datetime.now().strftime("%d/%m/%Y"))
                break
    for row in ws.iter_rows():
        codi_cell = row[0]
        if codi_cell.value is None:
            continue
        codi_str = str(codi_cell.value).replace('.0','').strip()
        if codi_str in stock.get(servei, {}):
            entrada = stock[servei][codi_str]
            ws.cell(row=codi_cell.row, column=5, value=entrada["comanda"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nom_servei = servei.capitalize()
    filename = f"Inventari_{nom_servei}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

if __name__ == "__main__":
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
    except:
        ip = "127.0.0.1"
    print(f"\n{'='*45}")
    print(f"  🏥  INVENTARI BST")
    print(f"{'='*45}")
    print(f"  App:          http://{ip}:8080/")
    print(f"  Transfusions: http://{ip}:8080/inventari/transfusions")
    print(f"  Immuno:       http://{ip}:8080/inventari/immuno")
    print(f"{'='*45}\n")
    app.run(host="0.0.0.0", port=8080, debug=True)
